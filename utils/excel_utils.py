#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/3/3 21:56
# @Author  : Gavin
# @File    : excel_utils.py


import xlrd
import xlwt

from openpyxl import Workbook, load_workbook


MAX_SHEET_ROWS = 65535
MAX_COLUMN_LENGTH = 30000


class HSSFUtils(object):
    """.xls文件"""
    def __init__(self, **kwargs):
        self.max_rows = kwargs.get("max_rows") if kwargs.get("max_rows") else MAX_SHEET_ROWS
        self.column_length = kwargs.get("column_length") if kwargs.get("column_length") else MAX_COLUMN_LENGTH

    def write_all(self, path, final_data):
        """
        写入excel.xls
        :param str path: 写入文件的路径
        :param dict final_data:
            {
                "sheet1": [
                    [column1, column2, column3]
                    [column1, column2, column3]
                ]
            }
        :return:
        """
        workbook = xlwt.Workbook()  # 新建一个工作簿
        for sheet_name, value in final_data.items():
            rows = len(value)  # 获取需要写入数据的行数
            if rows > self.max_rows:
                raise ValueError("max allowed rows: %d" % self.max_rows)
            columns = len(value[0])
            if columns > self.column_length:
                raise ValueError("max allowed columns: %d" % self.column_length)
            sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
            dt = None
            ro = 0
            col = 0
            try:
                for i in range(0, rows):
                    for j in range(0, columns):
                        data = value[i][j]
                        dt = data
                        ro = i
                        col = j
                        # write data to this sheet
                        sheet.write(i, j, data)
            except Exception as err:
                raise ValueError("data: %s, row: %s, column: %s, err info: %s" % (dt, ro, col, err))
        # save this workbook
        workbook.save(path)

    def read_all(self, path):
        """
        read all
        :param str path: .xls file path
        :return:
        """
        workbook = xlrd.open_workbook(path)
        sheets = workbook.sheet_names()
        workbook_dict = {}
        for sheet in sheets:
            self.read_single_sheet(workbook, sheet, workbook_dict)
        return workbook_dict

    def read_by_sheet(self, path, sheet):
        """
        read by sheet name
        :param str path: .xls file path
        :param str sheet: sheet name
        :return: sheet data
        """
        workbook = xlrd.open_workbook(path)
        workbook_dict = {}
        self.read_single_sheet(workbook, sheet, workbook_dict)
        return workbook_dict

    def read_single_sheet(self, workbook, sheet, workbook_dict):
        """
        read data from a sheet
        :param xlrd.Book workbook:
        :param str sheet: sheet name
        :param dict workbook_dict: data
        :return:
        """
        worksheet = workbook.sheet_by_name(sheet)
        sheet_list = []
        for i in range(0, worksheet.nrows):
            row_list = []
            for j in range(0, worksheet.ncols):
                cell_val = worksheet.cell_value(i, j)
                row_list.append(cell_val)
            sheet_list.append(row_list)
        workbook_dict[sheet] = sheet_list


class XSSFUtils(object):
    """.xlsx文件"""

    def read_by_sheet(self, path, sheet):
        workbook = load_workbook(path)
        sheet_obj = workbook[sheet]
        sheet_values = self.read_single_sheet(sheet_obj)
        return {sheet: sheet_values}

    def read_all(self, path):
        workbook = load_workbook(path)
        workbook_values = {}
        sheetnames = workbook.sheetnames
        for sheetname in sheetnames:
            sheet_obj = workbook[sheetname]
            sheet_values = self.read_single_sheet(sheet_obj)
            workbook_values[sheetname] = sheet_values
        return workbook_values

    def write_all(self, path, data):
        wb = Workbook()
        for sheetname, values in data.items():
            rows = len(values)
            columns = len(values[0])
            ws = wb.active
            ws.title = sheetname
            for i in range(0, rows):
                for j in range(0, columns):
                    ws.cell(i + 1, j + 1, values[i][j])
        wb.save(filename=path)

    def read_single_sheet(self, sheet_obj):
        sheet_list = []
        for row in sheet_obj.rows:
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            sheet_list.append(row_list)
        return sheet_list


class PandasExcel(object):
    # fixme
    pass
